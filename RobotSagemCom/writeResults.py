import time
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.worksheet.dimensions  import Dimension
from openpyxl import load_workbook , Workbook


workbook    =   None
wb          =   None
sheet       =   None
thin_border =   None
defaultFill =   None
greenFill   =   None
bold_font   =   None
     

def initialiseXlsxFile(xlsxfile):
    global workbook , sheet , thin_border , defaultFill , greenFill , bold_font
    try:
        workbook = load_workbook(filename=xlsxfile)
        print ("Open an existing file")
    except:
        print ("create a new file")
        workbook = Workbook()

    # Open Sheet if exist else Create a new Sheet
    try:
        sheet = workbook.active
    except :
        exit(0)
    
    #sheet = workbook.create_sheet(sheetName)
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    defaultFill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    greenFill = PatternFill(start_color='0cff0c',end_color='0cff0c',fill_type='solid')

    #Set header Style
    for header in sheet['A1:F1']:
        for col in header:
            col.fill    = defaultFill
            col.font    = bold_font
    #Set Border style
    for header in sheet['A1:F1']:
        for col in header:
            col.border = thin_border


    for header in sheet['A2:F2']:
        for col in header:
            col.border = thin_border

    
def saveFile(xlsxfile , sheetFile):
    global workbook
    adjustCellSize(sheetFile)
    workbook.save(filename=xlsxfile)
    adjustCellSize(sheetFile)
    

def adjustCellSize(sheetFile):
    for col in sheetFile['A:F']:
        if col[0].value != None:
            sheetFile.column_dimensions[col[0].column_letter].width=len(col[0].value)+8
            

def getNoneKey(cloNone , sheeyfile):
    val = 1
    maxLen = 1
    while True:
        celval = sheeyfile[cloNone+str(val)].value
        if celval == None:
            break
        else:
            if len(celval) > maxLen:
                maxLen = len(celval)
        val +=1
    return val , maxLen


def AdjustColones(cloNone , sheetFile):
    celval , maxLenght = getNoneKey(cloNone , sheetFile)
    for x in range (1 ,celval):
        celName =   cloNone+str(x)
        sheetFile.column_dimensions[sheetFile[celName].column_letter].width =maxLenght+8
        print(sheetFile.column_dimensions[sheetFile[celName].column_letter].width)
        print (x)



def insertWithBorder(xlsxfile , data = "" , name=None , rowNbr=None):
    global sheet , thin_border 
    if rowNbr != None:
        #sheet["A2"] = data
        sheet[name+str(rowNbr)] = str(data)
        sheet[name+str(rowNbr)].border = thin_border
    else:
        sheet[name] = str(data)
        sheet[name].border = thin_border

def writeinXlsxfile(xlsxfile):
    global workbook , sheet , thin_border , defaultFill , greenFill , bold_font
    initialiseXlsxFile(xlsxfile )
    initTestSuiteHeadSheet(xlsxfile )
    adjustCellSize(sheet)

    #insertSameData
    sheet["A2"] = "hello"
    #Save File / Modification
    #sheet(row=10 ,column=10 ).value= "hello"
    #Save File / Modification
    insertWithBorder(xlsxfile , data="test dta" ,name='A' ,rowNbr=5)
    saveFile(xlsxfile , sheet)




def initTestSuiteHeadSheet(xlsxfile):
    global workbook , sheet , thin_border , defaultFill , greenFill , bold_font
    sheet['A1']="TestSuite"
    sheet['B1']="TestCase"
    sheet['C1']="Summary"
    #sheet['D1']="Preconditions"
    sheet['D1']="Step"
    #sheet['F1']="ExpectedResults"
    sheet['E1']="Result"
    sheet['F1']="Comments"
    
   							

def writeTestSuiteInfo(xlsxfile , TestSuite=None , TestCase=None , Summary=None , Step=None 	, Result=None,	Comments=None):
    global workbook , sheet , thin_border , defaultFill , greenFill , bold_font
    initialiseXlsxFile(xlsxfile)
    initTestSuiteHeadSheet(xlsxfile)

    print(getNoneKey('D' , sheet))
    AdjustColones('D' , sheet)
    colNumnber = 1
    while True:
        if sheet['E'+str(colNumnber)].value == None:
            break
        colNumnber +=1
    
    if TestSuite != None:
        insertWithBorder(xlsxfile , data = TestSuite , name='A' , rowNbr=colNumnber)
        saveFile(xlsxfile , sheet)

    if TestCase != None:
        insertWithBorder(xlsxfile , data = TestCase , name='B' , rowNbr=colNumnber)
        saveFile(xlsxfile , sheet)

    if Summary != None:
        cellValue   =   sheet['D'+str(colNumnber)].value
        if cellValue != None:
            Summary += str(cellValue)
        insertWithBorder(xlsxfile , data =Summary , name='C' , rowNbr=colNumnber)
        saveFile(xlsxfile , sheet)
        
    if Step != None:
        cellValue   =   sheet['D'+str(colNumnber)].value
        if cellValue != None:
            Step += str(cellValue)
        insertWithBorder(xlsxfile , data = Step , name='D' , rowNbr=colNumnber)
        saveFile(xlsxfile , sheet)
    
    if Result != None:
        insertWithBorder(xlsxfile , data = Result , name='E' ,     rowNbr=colNumnber)
        saveFile(xlsxfile , sheet)
    
    if Comments != None:
        insertWithBorder(xlsxfile , data = Comments , name='F' ,   rowNbr=colNumnber)
        saveFile(xlsxfile , sheet)
    AdjustColones('D' , sheet)    
    saveFile(xlsxfile , sheet)
    AdjustColones('D' , sheet)

    

if __name__ == '__main__':
    #writeinXlsxfile("test.xlsx")
    fileTest    =   "testLibrary/pythonTest.xlsx"
    writeTestSuiteInfo(xlsxfile=fileTest , TestSuite="FU" , TestCase="Case" , Summary="None" , Step="None, reyfeyrfr ,frefrfre,ferfrefref,frefref" , Result="ok")