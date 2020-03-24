######################################################################
#######         Developper : Mansouri WISSEM
#######         This script is developped to used in robot framework script
#######         To create and insert data into an xlsx file
#######         There is many way to use this script
#######         
######################################################################

import time
import  os
import  sys
from openpyxl.styles import Color, PatternFill, Font, Border, Side , Alignment
from openpyxl.worksheet.dimensions  import Dimension
from openpyxl import load_workbook , Workbook
import string

class Results():
    # Initialise the class
    def __init__(self , fileName=None , sheetName=None):
        # initialise same data
        self.fileName   =   fileName
        self.sheetName  =   sheetName
        # config sheet file :
        # add same style
        self.bold_font      = Font(bold=True)
        self.thin_border    = Border(left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
            )
        self.defaultFill    = PatternFill(start_color='d3d3d3', 
            end_color='d3d3d3', 
            fill_type='solid'
            )
        self.greenFill      = PatternFill(start_color='0cff0c',
            end_color='0cff0c',
            fill_type='solid'
            )

        self.redFill        = PatternFill(start_color='FF0000',
            end_color='FF4500',
            fill_type='solid'
            )

        self.center_aligned_text    = Alignment(horizontal="center")
        self.listAlpha              =   string.ascii_uppercase
        self.numberList             =   ['0' , '1' , '2' , '3' , '4' , '5','6', '7' , '8' , '9']
        self.listPassTest           =   ['OK' , 'PASS']
        self.listFailTest           =   ['NOK' , 'FAIL']
    #create or open file
    def create_file(self):
        # This function used to 
        # 1 : Create a new xlsx file if dosen't exist
        # 2 : Initialise sheet 
        
        #if file exist open it
        if os.path.isfile(self.fileName):
            print ("Open an Existing File")
            myBook = load_workbook(filename=self.fileName)
        # create new file
        else:
            print ("Create a new file")
            myBook = Workbook()
        #Get all sheet in the file
        sheetsNames =   myBook.sheetnames
        # crate the specific sheet file
        if self.sheetName == None:
            mySheet = myBook.create_sheet("Results")
        # Test if sheet in argv
        else:
            #test if sheet exist
            #open the existing sheet
            if self.sheetName in sheetsNames:
                print ("open existing sheet")
                mySheet = myBook[self.sheetName]
            #create a new sheet file
            else:
                mySheet = myBook.create_sheet(self.sheetName)
        #Delete default sheet name
        self.remove_sheet_by_name("Sheet" , myBook)
        # Create Sheet And Workbook
        self.mySheet    =   mySheet
        self.myBook     =   myBook
        #Save file
        self.save_data_in_file()

    # Function Save file
    def save_data_in_file(self):
        self.myBook.save(filename=self.fileName)

    # Function to delete a specific sheet in xlsx file
    def remove_sheet_by_name(self , sheetFineName , workbookName):
        sheetsNames =   workbookName.sheetnames
        if sheetFineName in sheetsNames:
            print ("Delete Sheet")
            workbookName.remove(workbookName[sheetFineName])

    # This function used to adjust a specific cell (col) in th e active sheet file
    def adjust_cell(self , cellName , cellSize):
        if self.mySheet[cellName] != None:
            self.mySheet.column_dimensions[self.mySheet[cellName].column_letter].width=cellSize
            self.save_data_in_file()

    # This function used to adjust the size off all sheet cell by the headre
    def adjust_cells_size(self, *listArgs):
        # How to use it
        # Pass all coll with the specific size
        # example of usage:
        # 1 : adjust_cells_size('A:20' , 'B:20' ,'C:30')
        # 1 : adjust_cells_size('0:20' , '1:20' ,'2:30')
        for x in listArgs:
            cellName    =   x.split(':')
            col         =   (cellName[0].upper()).strip()
            if col in self.listAlpha:
                index   =   col+"1"
            else:
                index   =   self.listAlpha[int(col)]+"1"
            size        =   int(cellName[1].strip())
            self.adjust_cell(index , size)

    # this function used to create the sheet header
    def create_header_file(self, *argsList):
        # Example usage: create_header_file("TestSuite" , "Summary"  , "Test Case" , "Results")
        key         =   0
        for value in argsList:
            strKey     =   self.listAlpha[key]+"1"
            self.insert_cell_header_data(strKey , value)
            key +=1

    # insert data in header cell
    def insert_cell_header_data(self, cellName , data):
        # There is a 3 methods to call this function
        # 1 :using full cell name example:   insert_cell_header_data("A1" , "data")
        # 2 :using first string col name example:   insert_cell_header_data("A" , "data")
        # 3 :using int cell name example:   insert_cell_header_data(1 , "data")
        strCelName  =   str(cellName)
        if len(strCelName) == 1 and strCelName[0] in self.listAlpha:
            cellNameKey =   strCelName+"1"
        elif strCelName[0] in self.listAlpha and strCelName[1] in self.numberList:
            cellNameKey =   strCelName
        else:
            key         =   cellName
            cellNameKey =   self.listAlpha[key]+"1"
            print (cellNameKey)
        self.mySheet[cellNameKey]              =   data
        self.mySheet[cellNameKey].border       =   self.thin_border
        self.mySheet[cellNameKey].fill         =   self.defaultFill
        self.mySheet[cellNameKey].font         =   self.bold_font
        self.mySheet[cellNameKey].alignment    =   self.center_aligned_text
        self.save_data_in_file()
        
    # insert data in specific cell
    def insert_cell_data(self, cellName , data):
        # insert data in specific cell
        # The cellName can be
        # complete name string  : like A1 , B5 
        # row : col : like      1:5         22:14
        # Example usage : ResultsObject.insert_cell_data("A5" , "OK")
        #                 ResultsObject.insert_cell_data("1:5" , "OK")
        # import : in the second way the numbere begin from 0
        strCellName =   str(cellName)
        if strCellName[0] in self.listAlpha:
            self.mySheet[strCellName]       =   data
        else:
            arrKey                          =   strCellName.split(':')
            firstKey                        =   int(arrKey[0])
            secondKey                       =   int(arrKey[1]) + 1
            cellData                        =   self.listAlpha[firstKey]+str(secondKey)
            self.mySheet[cellData]          =   data
            self.mySheet[cellData].border   =   self.thin_border
        self.save_data_in_file()

    # insert a data by col name or col value
    def insert_col_data(self, cellName , data , appendData="False"):
        # insert a data by col name or col value
        # The cellName can be
        # string  : like A , B
        # integer : 1  ,   14 , 0
        # Example usage : ResultsObject.insert_col_data("A" , "OK")
        #                 ResultsObject.insert_col_data("1" , "OK")
        # import : in the second way the numbere begin from 0
        strCellName =   str(cellName)
        if strCellName in self.listAlpha:
            key             =   self.get_empty_cell(strCellName)
        else:
            strCellName     =   self.listAlpha[int(cellName)]
            key             =   self.get_empty_cell(strCellName)
        self.mySheet[key]   =   data
        self.mySheet[key].border   =   self.thin_border
        self.save_data_in_file()

    # This function used to get the next cell to insert data
    def get_empty_cell(self, cel):
        strCel  =   str(cel)
        Testkey     =   1
        if len(strCel) == 1  and strCel in self.listAlpha:
            celName =   strCel
        elif strCel[0] in self.listAlpha and strCel[1] in self.numberList:
            celName =   strCel[0]
        else:
            intCel  =   int(cel)
            celName =   self.listAlpha[intCel]
            celName =   celName
        
        while True:
            celNameValue =   celName+str(Testkey)
            if self.mySheet[celNameValue].value == None:
                break
            Testkey +=1
        return celNameValue

    #insert data by header name
    def insert_colName_data(self, indexColName , data):
        #This function used to insert data into a col by the name of the head col
        #For example inset data into a col Results
        testInsert  =   False
        if len(indexColName) == 1:
            x           =   indexColName
            testInsert  =   True
        else:
            # Get cell by header name
            for x in self.listAlpha:
                colName     =   x+"1"
                if (self.mySheet[colName].value).lower() == indexColName.lower():
                    testInsert  =   True
                    break
        if testInsert == True:
            self.insert_col_data(x , data)    

    def insert_col_result(self, indexColName , data):
        #This function used to insert data into a col result
        #inset data into a col Results
        # Usage:
        # 1 : insert_col_result(col_Result_Name , data) example :  insert_col_result('Results' , 'OK')
        # 2 : insert_col_result(col_index_Name , data) example  insert_col_result('C' , 'OK')
        # 2 : insert_col_result(col_index_value , data) example  insert_col_result(3 , 'OK')
        testInsert  =   False
        indexColName    =   str(indexColName)

        if len(indexColName) == 1:
            if indexColName in self.listAlpha:
                x           =   indexColName
            else:
                x           =   self.listAlpha[int(indexColName)]
            testInsert  =   True
        else:
            # Get cell by header name
            for x in self.listAlpha:
                colName     =   x+"1"
                if (self.mySheet[colName].value).lower() == indexColName.lower():
                    testInsert  =   True
                    break
        celKey  =   self.get_empty_cell(x)
        if testInsert == True:
            if data.upper() in self.listPassTest :
                self.mySheet[celKey]            =   "OK"
                self.mySheet[celKey].fill       =   self.greenFill
            elif data in self.listFailTest :
                self.mySheet[celKey]            =   "NOK"
                self.mySheet[celKey].fill       =   self.redFill
            self.mySheet[celKey].border         =   self.thin_border    
        self.save_data_in_file()


    def insert_row_data(self, *argListData):
        # Insert row (many) data 
        # usage:
        # 1 : insert_row_data('data 1' , 'data 2' , 'data 3')
        # 2 : insert_row_data('A:data 1' , 'B:data 2' , 'C:data 3')
        # 3 : insert_row_data('0:data 1' , '1:data 2' , '2:data 3')
        compt = 0
        for target_list in argListData:
            if target_list.find(':') != -1:
                target_list     =   target_list.split(':')
                key     =   target_list[0].strip()
                if key in self.listAlpha:
                    index = key
                else:
                    index = self.listAlpha[int(key)]
                data    =   target_list[1]
                print (index , data)
                self.insert_col_data(index , data)
            else :
                data    =   target_list
                index   =   self.listAlpha[compt]
                print (index , data)
                self.insert_col_data(index , data)
                compt   +=1
                

if __name__ == "__main__":
    currentPath     =   os.getcwd()
    currentPath     =   currentPath.replace('\\' , '/')
    print (currentPath)

    pathFolder      =   currentPath+"/RobotSagemCom/testLibrary/"
    firstFile       =   pathFolder+"testInPythonScript.xlsx"
    firstResults    =   Results(firstFile , "Results")
    firstResults.create_file()